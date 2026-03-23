from __future__ import annotations

import shutil
from copy import copy
from pathlib import Path

from openpyxl import load_workbook

from journal_tracker.discover import (
    SUGGESTIONS_SHEET,
    SourceCandidate,
    discover_journals,
)


def test_discover_journals_writes_suggestion_sheet_for_missing_entries(tmp_path: Path) -> None:
    source_workbook = Path("examples/turkish_politics_articles_database.sample.xlsx")
    workbook_path = tmp_path / "tracker.xlsx"
    shutil.copy2(source_workbook, workbook_path)

    workbook = load_workbook(workbook_path)
    directory = workbook["Journal Directory"]
    next_row = directory.max_row + 1
    for column_index in range(1, directory.max_column + 1):
        source = directory.cell(row=2, column=column_index)
        target = directory.cell(row=next_row, column=column_index)
        if source.has_style:
            target._style = source._style
        target.font = copy(source.font)
        target.fill = copy(source.fill)
        target.border = copy(source.border)
        target.alignment = copy(source.alignment)
        target.protection = copy(source.protection)
        target.number_format = source.number_format
    directory.cell(row=next_row, column=1).value = "Imaginary Politics Review"
    directory.cell(row=next_row, column=2).value = "Demo Press"
    workbook.save(workbook_path)
    workbook.close()

    seen_queries: list[tuple[str, str]] = []

    def fake_searcher(journal_name: str, publisher: str) -> list[SourceCandidate]:
        seen_queries.append((journal_name, publisher))
        return [
            SourceCandidate(
                source_id="https://openalex.org/S123",
                display_name="Imaginary Politics Review",
                publisher="Demo Press",
                issn_l="1234-5678",
                works_count=150,
                confidence_note="exact title match; publisher match",
            ),
            SourceCandidate(
                source_id="https://openalex.org/S456",
                display_name="Imaginary Politics Quarterly",
                publisher="Demo Press",
                issn_l="9876-5432",
                works_count=90,
                confidence_note="partial title match",
            ),
        ]

    summary = discover_journals(
        workbook_path=workbook_path,
        source_searcher=fake_searcher,
    )

    assert summary.missing_journals == 1
    assert summary.suggestion_rows == 1
    assert seen_queries == [("Imaginary Politics Review", "Demo Press")]

    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    assert SUGGESTIONS_SHEET in workbook.sheetnames
    sheet = workbook[SUGGESTIONS_SHEET]
    assert sheet["A1"].value == "Journal Name"
    assert sheet["A2"].value == "Imaginary Politics Review"
    assert sheet["C2"].value == "https://openalex.org/S123"
    assert sheet["D2"].value == "Imaginary Politics Review"
    assert sheet["E2"].value == "Demo Press"
    assert sheet["G2"].value == "Imaginary Politics Quarterly | https://openalex.org/S456"
    assert sheet["I2"].value == "exact title match; publisher match"
    workbook.close()


def test_discover_journals_writes_header_only_when_everything_is_mapped(tmp_path: Path) -> None:
    source_workbook = Path("examples/turkish_politics_articles_database.sample.xlsx")
    workbook_path = tmp_path / "tracker.xlsx"
    shutil.copy2(source_workbook, workbook_path)

    summary = discover_journals(
        workbook_path=workbook_path,
        source_searcher=lambda journal_name, publisher: [],
    )

    assert summary.missing_journals == 0
    assert summary.suggestion_rows == 0

    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    sheet = workbook[SUGGESTIONS_SHEET]
    assert sheet.max_row == 1
    workbook.close()
