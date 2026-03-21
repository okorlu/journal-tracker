from __future__ import annotations

import csv
import json
import os
import shutil
from datetime import date
from pathlib import Path

from openpyxl import load_workbook

from journal_tracker.cli import env_file_candidates, load_runtime_env, resolve_run_options
from journal_tracker.profiles import load_profile
from journal_tracker.sync import (
    META_SHEET,
    default_config_path,
    export_articles_to_csv,
    sync_workbook,
)


def test_sync_workbook_is_idempotent_with_sample_workbook(tmp_path: Path) -> None:
    source_workbook = Path("examples/turkish_politics_articles_database.sample.xlsx")
    workbook_path = tmp_path / "tracker.xlsx"
    shutil.copy2(source_workbook, workbook_path)

    payloads = {
        "https://openalex.org/S77485876": [
            {
                "id": "https://openalex.org/W-existing",
                "display_name": "Sample Existing Article",
                "publication_year": 2025,
                "publication_date": "2025-02-01",
                "doi": "https://doi.org/10.5555/sample-existing",
                "biblio": {"volume": "25", "issue": "1", "first_page": "1", "last_page": "12"},
                "authorships": [{"author": {"display_name": "Aylin Demo"}}],
                "keywords": [{"display_name": "Existing Topic"}],
            },
            {
                "id": "https://openalex.org/W-new-1",
                "display_name": "Fresh Turkish Studies Article",
                "publication_year": 2026,
                "publication_date": "2026-03-01",
                "doi": "https://doi.org/10.5555/fresh-ts",
                "biblio": {"volume": "26", "issue": "1", "first_page": "13", "last_page": "28"},
                "authorships": [{"author": {"display_name": "Mert Example"}}],
                "keywords": [{"display_name": "Political science"}],
            },
        ],
        "https://openalex.org/S6959732": [
            {
                "id": "https://openalex.org/W-new-2",
                "display_name": "Fresh Party Politics Article",
                "publication_year": 2026,
                "publication_date": "2026-03-02",
                "doi": None,
                "primary_location": {"landing_page_url": "https://example.org/party-politics"},
                "biblio": {"volume": "32", "issue": "2", "first_page": "44", "last_page": "58"},
                "authorships": [{"author": {"display_name": "Deniz Example"}}],
                "topics": [{"display_name": "Party systems"}],
            }
        ],
    }

    def fake_fetcher(source_id: str, cutoff_date: date, api_key: str):
        assert api_key == "test-key"
        assert cutoff_date == date(2023, 3, 15)
        return payloads.get(source_id, [])

    summary = sync_workbook(
        workbook_path=workbook_path,
        config_path=default_config_path(),
        api_key="test-key",
        years=3,
        dry_run=False,
        today=date(2026, 3, 15),
        fetcher=fake_fetcher,
    )

    assert summary.total_new_rows == 2
    assert summary.total_duplicates == 1
    assert summary.backup_path is not None
    assert summary.backup_path.exists()

    workbook = load_workbook(workbook_path)
    sheet = workbook["Articles"]
    assert sheet.max_row == 5
    assert sheet.max_column == 11
    assert sheet["G1"].value == "DOI"
    assert sheet["H1"].value == "Article URL"
    assert sheet["K1"].value == "Added At"
    assert sheet["A4"].value == "Fresh Turkish Studies Article"
    assert sheet["A5"].value == "Fresh Party Politics Article"
    assert sheet["G4"].value == "https://doi.org/10.5555/fresh-ts"
    assert sheet["H4"].value == "https://openalex.org/W-new-1"
    assert sheet["H5"].value == "https://example.org/party-politics"
    assert sheet["K4"].value
    assert sheet["K5"].value
    assert workbook[META_SHEET].sheet_state == "hidden"
    workbook.close()

    second_summary = sync_workbook(
        workbook_path=workbook_path,
        config_path=default_config_path(),
        api_key="test-key",
        years=3,
        dry_run=False,
        today=date(2026, 3, 15),
        fetcher=fake_fetcher,
    )

    assert second_summary.total_new_rows == 0
    assert second_summary.backup_path is None


def test_export_articles_to_csv_writes_current_articles_sheet(tmp_path: Path) -> None:
    source_workbook = Path("examples/turkish_politics_articles_database.sample.xlsx")
    workbook_path = tmp_path / "tracker.xlsx"
    csv_path = tmp_path / "exports" / "tracker.csv"
    shutil.copy2(source_workbook, workbook_path)

    export_articles_to_csv(workbook_path, csv_path)

    assert csv_path.exists()
    with csv_path.open(encoding="utf-8", newline="") as handle:
        rows = list(csv.reader(handle))

    assert rows[0] == [
        "Article Title",
        "Author(s)",
        "Journal",
        "Volume/Issue",
        "Year",
        "Pages",
        "DOI",
        "Article URL",
        "Cluster",
        "Key Topics",
        "Added At",
    ]
    assert rows[1][0] == "Sample Existing Article"
    assert rows[1][6] == "https://doi.org/10.5555/sample-existing"
    assert rows[1][7] == ""
    assert rows[1][10] == ""


def test_sync_workbook_can_filter_to_profile_journal_subset(tmp_path: Path) -> None:
    source_workbook = Path("examples/turkish_politics_articles_database.sample.xlsx")
    workbook_path = tmp_path / "tracker.xlsx"
    shutil.copy2(source_workbook, workbook_path)

    payloads = {
        "https://openalex.org/S6959732": [
            {
                "id": "https://openalex.org/W-new-party",
                "display_name": "Profile Limited Party Article",
                "publication_year": 2026,
                "publication_date": "2026-03-02",
                "doi": "https://doi.org/10.5555/profile-party",
                "biblio": {"volume": "32", "issue": "2", "first_page": "44", "last_page": "58"},
                "authorships": [{"author": {"display_name": "Deniz Example"}}],
                "topics": [{"display_name": "Party systems"}],
            }
        ]
    }

    def fake_fetcher(source_id: str, cutoff_date: date, api_key: str):
        assert api_key == "test-key"
        assert cutoff_date == date(2023, 3, 15)
        return payloads.get(source_id, [])

    summary = sync_workbook(
        workbook_path=workbook_path,
        config_path=default_config_path(),
        api_key="test-key",
        years=3,
        dry_run=False,
        journal_names=["Party Politics"],
        today=date(2026, 3, 15),
        fetcher=fake_fetcher,
    )

    assert [result.journal_name for result in summary.journal_results] == ["Party Politics"]
    assert summary.total_new_rows == 1

    workbook = load_workbook(workbook_path)
    sheet = workbook["Articles"]
    assert sheet.max_row == 4
    assert sheet["H1"].value == "Article URL"
    assert sheet["K1"].value == "Added At"
    assert sheet["A4"].value == "Profile Limited Party Article"
    assert sheet["G4"].value == "https://doi.org/10.5555/profile-party"
    assert sheet["H4"].value == "https://openalex.org/W-new-party"
    assert sheet["K4"].value
    workbook.close()


def test_profile_can_supply_paths_and_defaults(tmp_path: Path) -> None:
    workbook_path = tmp_path / "tracker.xlsx"
    profile_path = tmp_path / "starter.json"
    workbook_path.write_text("", encoding="utf-8")
    profile_path.write_text(
        json.dumps(
            {
                "workbook": "tracker.xlsx",
                "csv_output": "exports/tracker.csv",
                "config": str(default_config_path()),
                "years": 4,
                "articles_sheet": "Articles",
                "directory_sheet": "Journal Directory",
                "journals": ["Party Politics", "Turkish Studies"],
            }
        ),
        encoding="utf-8",
    )

    args = type(
        "Args",
        (),
        {
            "profile": str(profile_path),
            "workbook": None,
            "config": None,
            "years": None,
            "dry_run": True,
            "csv_output": None,
        },
    )()

    profile, options = resolve_run_options(args)

    assert profile is not None
    assert profile == load_profile(profile_path)
    assert options["workbook_path"] == workbook_path.resolve()
    assert options["config_path"] == default_config_path()
    assert options["years"] == 4
    assert options["articles_sheet"] == "Articles"
    assert options["directory_sheet"] == "Journal Directory"
    assert options["journal_names"] == ("Party Politics", "Turkish Studies")
    assert options["csv_output_path"] == (tmp_path / "exports" / "tracker.csv").resolve()


def test_resolve_run_options_normalizes_direct_cli_paths(tmp_path: Path, monkeypatch) -> None:
    workspace = tmp_path / "workspace"
    workspace.mkdir()
    workbook_path = workspace / "tracker.xlsx"
    config_path = workspace / "sources.json"
    csv_path = workspace / "exports" / "tracker.csv"
    workbook_path.write_text("", encoding="utf-8")
    config_path.write_text("[]", encoding="utf-8")
    csv_path.parent.mkdir()

    home_dir = tmp_path / "home"
    home_dir.mkdir()
    home_dir_str = str(home_dir)
    monkeypatch.setenv("HOME", str(home_dir))
    monkeypatch.setenv("USERPROFILE", home_dir_str)
    monkeypatch.setenv("HOMEDRIVE", home_dir.drive or "")
    monkeypatch.setenv(
        "HOMEPATH",
        home_dir_str[len(home_dir.drive) :] if home_dir.drive else home_dir_str,
    )

    relative_workbook = Path("tracker.xlsx")
    relative_config = Path("sources.json")
    tilde_csv = "~/journal-tracker.csv"

    args = type(
        "Args",
        (),
        {
            "profile": None,
            "workbook": str(relative_workbook),
            "config": str(relative_config),
            "years": None,
            "dry_run": True,
            "csv_output": tilde_csv,
        },
    )()

    original_cwd = Path.cwd()
    try:
        os.chdir(workspace)
        _, options = resolve_run_options(args)
    finally:
        os.chdir(original_cwd)

    assert options["workbook_path"] == workbook_path.resolve()
    assert options["config_path"] == config_path.resolve()
    assert options["csv_output_path"] == (home_dir / "journal-tracker.csv").resolve()


def test_env_file_candidates_prefer_run_contexts_and_dedupe(tmp_path: Path) -> None:
    current_dir = tmp_path / "cwd"
    profile_dir = tmp_path / "profiles"
    workbook_dir = tmp_path / "data"
    for directory in (current_dir, profile_dir, workbook_dir):
        directory.mkdir()

    profile_path = profile_dir / "starter.json"
    profile_path.write_text(json.dumps({"workbook": "../data/tracker.xlsx"}), encoding="utf-8")
    profile = load_profile(profile_path)
    assert profile.workbook_path is not None

    candidates = env_file_candidates(current_dir, profile, profile.workbook_path)

    assert candidates[0] == (profile_dir / ".env").resolve()
    assert candidates[1] == (workbook_dir / ".env").resolve()
    assert candidates[-1] == (current_dir / ".env").resolve()
    assert len(candidates) == len(set(candidates))


def test_load_runtime_env_finds_repo_env_outside_repo_root(tmp_path: Path, monkeypatch) -> None:
    workspace = tmp_path / "workspace"
    workspace.mkdir()
    workbook_path = workspace / "tracker.xlsx"
    workbook_path.write_text("", encoding="utf-8")

    original_cwd = Path.cwd()
    repo_env_path = Path("src/journal_tracker/cli.py").resolve().parents[2] / ".env"
    original_repo_env = (
        repo_env_path.read_text(encoding="utf-8") if repo_env_path.exists() else None
    )
    monkeypatch.delenv("OPENALEX_API_KEY", raising=False)

    try:
        os.chdir(tmp_path)
        repo_env_path.write_text("OPENALEX_API_KEY=repo-test-key\n", encoding="utf-8")
        load_runtime_env(None, workbook_path)
    finally:
        os.chdir(original_cwd)
        if original_repo_env is None:
            repo_env_path.unlink(missing_ok=True)
        else:
            repo_env_path.write_text(original_repo_env, encoding="utf-8")

    assert os.getenv("OPENALEX_API_KEY") == "repo-test-key"


def test_sync_workbook_emits_progress_messages(tmp_path: Path) -> None:
    source_workbook = Path("examples/turkish_politics_articles_database.sample.xlsx")
    workbook_path = tmp_path / "tracker.xlsx"
    shutil.copy2(source_workbook, workbook_path)
    messages: list[str] = []

    def fake_fetcher(source_id: str, cutoff_date: date, api_key: str):
        del source_id, cutoff_date, api_key
        return []

    summary = sync_workbook(
        workbook_path=workbook_path,
        config_path=default_config_path(),
        api_key="test-key",
        years=3,
        dry_run=True,
        journal_names=["Party Politics"],
        today=date(2026, 3, 15),
        fetcher=fake_fetcher,
        progress_callback=messages.append,
    )

    assert summary.total_new_rows == 0
    assert any("Loading journal config" in message for message in messages)
    assert any("Reading journal directory" in message for message in messages)
    assert any("Fetching Party Politics" in message for message in messages)
    assert any("Party Politics: fetched=0 new=0 duplicates=0" in message for message in messages)


def test_sync_workbook_can_fill_missing_doi_from_crossref(tmp_path: Path) -> None:
    source_workbook = Path("examples/turkish_politics_articles_database.sample.xlsx")
    workbook_path = tmp_path / "tracker.xlsx"
    shutil.copy2(source_workbook, workbook_path)

    payloads = {
        "https://openalex.org/S6959732": [
            {
                "id": "https://openalex.org/W-crossref",
                "display_name": "Crossref Rescued Article",
                "publication_year": 2026,
                "publication_date": "2026-03-02",
                "doi": None,
                "primary_location": {
                    "landing_page_url": "https://www.tandfonline.com/toc/fpal20/current"
                },
                "biblio": {"volume": "32", "issue": "2", "first_page": "44", "last_page": "58"},
                "authorships": [{"author": {"display_name": "Deniz Example"}}],
                "topics": [{"display_name": "Party systems"}],
            }
        ]
    }

    def fake_fetcher(source_id: str, cutoff_date: date, api_key: str):
        assert api_key == "test-key"
        assert cutoff_date == date(2023, 3, 15)
        return payloads.get(source_id, [])

    def fake_crossref_lookup(work, directory_entry):
        assert work["display_name"] == "Crossref Rescued Article"
        assert directory_entry.journal_name == "Party Politics"
        return type(
            "CrossrefCandidateStub",
            (),
            {
                "doi_url": "https://doi.org/10.5555/crossref-rescued",
                "article_url": "https://publisher.example.org/article/crossref-rescued",
            },
        )()

    summary = sync_workbook(
        workbook_path=workbook_path,
        config_path=default_config_path(),
        api_key="test-key",
        years=3,
        dry_run=False,
        journal_names=["Party Politics"],
        today=date(2026, 3, 15),
        fetcher=fake_fetcher,
        crossref_lookup=fake_crossref_lookup,
    )

    assert summary.total_new_rows == 1

    workbook = load_workbook(workbook_path)
    sheet = workbook["Articles"]
    assert sheet["G4"].value == "https://doi.org/10.5555/crossref-rescued"
    assert sheet["H4"].value == "https://publisher.example.org/article/crossref-rescued"
    workbook.close()
