from __future__ import annotations

import json
from copy import copy
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Callable, Iterable
from urllib.error import HTTPError, URLError
from urllib.parse import urlencode

from openpyxl import load_workbook

from journal_tracker.sync import (
    DIRECTORY_JOURNAL_HEADER,
    DIRECTORY_PUBLISHER_HEADER,
    DIRECTORY_SHEET,
    ProgressCallback,
    default_config_path,
    emit_progress,
    json_get,
    normalize_text,
)

OPENALEX_SOURCES_API_URL = "https://api.openalex.org/sources"
SUGGESTIONS_SHEET = "Journal Match Suggestions"
DEFAULT_SUGGESTION_LIMIT = 3


@dataclass(frozen=True)
class SourceCandidate:
    source_id: str
    display_name: str
    publisher: str
    issn_l: str
    works_count: int
    confidence_note: str


@dataclass(frozen=True)
class JournalSuggestion:
    journal_name: str
    status: str
    suggested_source_id: str
    suggested_display_name: str
    suggested_publisher: str
    suggested_issn_l: str
    candidate_2: str
    candidate_3: str
    confidence_note: str


@dataclass(frozen=True)
class DiscoverySummary:
    workbook_path: Path
    config_path: Path
    journals_checked: int
    missing_journals: int
    suggestion_rows: int
    output_sheet: str


def load_config_names(config_path: Path) -> set[str]:
    items = json.loads(config_path.read_text(encoding="utf-8"))
    return {item["journal_name"] for item in items}


def read_directory_rows(
    workbook_path: Path,
    directory_sheet_name: str = DIRECTORY_SHEET,
) -> list[tuple[str, str]]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    if directory_sheet_name not in workbook.sheetnames:
        workbook.close()
        raise ValueError(f"Workbook is missing the '{directory_sheet_name}' sheet.")

    sheet = workbook[directory_sheet_name]
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    header_indexes = {
        str(value or "").strip(): index
        for index, value in enumerate(header_row)
        if str(value or "").strip()
    }
    journal_index = header_indexes.get(DIRECTORY_JOURNAL_HEADER)
    if journal_index is None:
        workbook.close()
        raise ValueError(
            f"'{directory_sheet_name}' must include a '{DIRECTORY_JOURNAL_HEADER}' column."
        )
    publisher_index = header_indexes.get(DIRECTORY_PUBLISHER_HEADER)

    rows: list[tuple[str, str]] = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        journal_name = str(row[journal_index] or "").strip()
        if not journal_name:
            continue
        publisher = ""
        if publisher_index is not None and publisher_index < len(row):
            publisher = str(row[publisher_index] or "").strip()
        rows.append((journal_name, publisher))
    workbook.close()
    return rows


def openalex_source_get(url: str) -> dict[str, Any]:
    return json_get(url, "journal-tracker-source-discovery/1.0")


def candidate_confidence_note(
    journal_name: str,
    expected_publisher: str,
    candidate: dict[str, Any],
) -> str:
    notes: list[str] = []
    normalized_journal = normalize_text(journal_name)
    display_name = candidate.get("display_name") or ""
    normalized_display_name = normalize_text(display_name)
    alternate_titles = [
        normalize_text(value) for value in (candidate.get("alternate_titles") or [])
    ]
    publisher = (
        (candidate.get("host_organization_name") or "")
        or (candidate.get("summary_stats") or {}).get("publisher", "")
    )

    if normalized_display_name == normalized_journal:
        notes.append("exact title match")
    elif normalized_display_name.startswith(normalized_journal) or normalized_journal.startswith(
        normalized_display_name
    ):
        notes.append("close title match")
    elif normalized_journal in alternate_titles:
        notes.append("alternate title match")
    else:
        notes.append("partial title match")

    if expected_publisher and publisher:
        normalized_publisher = normalize_text(expected_publisher)
        normalized_candidate_publisher = normalize_text(publisher)
        if normalized_publisher and normalized_publisher == normalized_candidate_publisher:
            notes.append("publisher match")

    return "; ".join(notes)


def rank_source_candidates(
    journal_name: str,
    expected_publisher: str,
    results: Iterable[dict[str, Any]],
) -> list[SourceCandidate]:
    normalized_journal = normalize_text(journal_name)
    ranked: list[tuple[tuple[int, int, int], SourceCandidate]] = []
    for candidate in results:
        display_name = candidate.get("display_name") or ""
        normalized_display_name = normalize_text(display_name)
        alternate_titles = {
            normalize_text(value) for value in (candidate.get("alternate_titles") or []) if value
        }
        publisher = candidate.get("host_organization_name") or ""
        confidence_note = candidate_confidence_note(journal_name, expected_publisher, candidate)
        exact = int(normalized_display_name == normalized_journal)
        alternate = int(normalized_journal in alternate_titles)
        works_count = int(candidate.get("works_count") or 0)
        ranked.append(
            (
                (exact, alternate, works_count),
                SourceCandidate(
                    source_id=candidate.get("id") or "",
                    display_name=display_name,
                    publisher=publisher,
                    issn_l=candidate.get("issn_l") or "",
                    works_count=works_count,
                    confidence_note=confidence_note,
                ),
            )
        )

    ranked.sort(key=lambda item: item[0], reverse=True)
    return [candidate for _, candidate in ranked]


def search_openalex_sources(
    journal_name: str,
    expected_publisher: str,
    limit: int = DEFAULT_SUGGESTION_LIMIT,
    getter: Callable[[str], dict[str, Any]] = openalex_source_get,
) -> list[SourceCandidate]:
    params = {
        "search": journal_name,
        "per-page": str(max(limit, DEFAULT_SUGGESTION_LIMIT)),
        "select": "id,display_name,host_organization_name,issn_l,alternate_titles,works_count",
    }
    url = f"{OPENALEX_SOURCES_API_URL}?{urlencode(params)}"
    try:
        payload = getter(url)
    except HTTPError as exc:
        body = exc.read().decode("utf-8", errors="replace")
        raise RuntimeError(
            f"OpenAlex source search failed for '{journal_name}': HTTP {exc.code} {body}"
        ) from exc
    except URLError as exc:
        raise RuntimeError(f"OpenAlex source search failed for '{journal_name}': {exc}") from exc

    results = payload.get("results", [])
    return rank_source_candidates(journal_name, expected_publisher, results)[:limit]


def build_suggestion(
    journal_name: str,
    expected_publisher: str,
    candidates: list[SourceCandidate],
) -> JournalSuggestion:
    top = candidates[0] if candidates else None
    candidate_2 = ""
    candidate_3 = ""
    if len(candidates) > 1:
        candidate_2 = f"{candidates[1].display_name} | {candidates[1].source_id}"
    if len(candidates) > 2:
        candidate_3 = f"{candidates[2].display_name} | {candidates[2].source_id}"

    return JournalSuggestion(
        journal_name=journal_name,
        status="missing_config",
        suggested_source_id=top.source_id if top else "",
        suggested_display_name=top.display_name if top else "",
        suggested_publisher=top.publisher if top else expected_publisher,
        suggested_issn_l=top.issn_l if top else "",
        candidate_2=candidate_2,
        candidate_3=candidate_3,
        confidence_note=top.confidence_note if top else "no OpenAlex candidates found",
    )


def write_suggestions_sheet(
    workbook_path: Path,
    suggestions: list[JournalSuggestion],
    sheet_name: str = SUGGESTIONS_SHEET,
) -> None:
    workbook = load_workbook(workbook_path)
    directory = workbook[DIRECTORY_SHEET]
    if sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        sheet.delete_rows(1, sheet.max_row)
    else:
        insert_index = workbook.sheetnames.index(DIRECTORY_SHEET) + 1
        sheet = workbook.create_sheet(sheet_name, insert_index)

    header_source = directory["A1"]
    data_source = directory["A2"]
    headers = [
        "Journal Name",
        "Status",
        "Suggested Source ID",
        "Suggested Display Name",
        "Suggested Publisher",
        "Suggested ISSN-L",
        "Candidate 2",
        "Candidate 3",
        "Confidence Note",
        "Approve?",
    ]

    for column_index, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=column_index)
        cell.value = header
        if header_source.has_style:
            cell._style = copy(header_source._style)
        cell.font = copy(header_source.font)
        cell.fill = copy(header_source.fill)
        cell.border = copy(header_source.border)
        cell.alignment = copy(header_source.alignment)
        cell.protection = copy(header_source.protection)
        cell.number_format = header_source.number_format

    for row_index, suggestion in enumerate(suggestions, start=2):
        values = [
            suggestion.journal_name,
            suggestion.status,
            suggestion.suggested_source_id,
            suggestion.suggested_display_name,
            suggestion.suggested_publisher,
            suggestion.suggested_issn_l,
            suggestion.candidate_2,
            suggestion.candidate_3,
            suggestion.confidence_note,
            "",
        ]
        for column_index, value in enumerate(values, start=1):
            cell = sheet.cell(row=row_index, column=column_index)
            cell.value = value
            if data_source.has_style:
                cell._style = copy(data_source._style)
            cell.font = copy(data_source.font)
            cell.fill = copy(data_source.fill)
            cell.border = copy(data_source.border)
            cell.alignment = copy(data_source.alignment)
            cell.protection = copy(data_source.protection)
            cell.number_format = data_source.number_format
            if column_index == 3 and value:
                cell.hyperlink = value

    workbook.save(workbook_path)
    workbook.close()


def discover_journals(
    workbook_path: Path,
    config_path: Path | None = None,
    directory_sheet_name: str = DIRECTORY_SHEET,
    progress_callback: ProgressCallback | None = None,
    source_searcher: Callable[[str, str], list[SourceCandidate]] = search_openalex_sources,
) -> DiscoverySummary:
    workbook_path = workbook_path.expanduser().resolve()
    resolved_config_path = (config_path or default_config_path()).expanduser().resolve()

    if not workbook_path.exists():
        raise FileNotFoundError(f"Workbook not found: {workbook_path}")
    if not resolved_config_path.exists():
        raise FileNotFoundError(f"Config not found: {resolved_config_path}")

    emit_progress(progress_callback, f"Reading journals from {workbook_path}")
    directory_rows = read_directory_rows(workbook_path, directory_sheet_name)
    known_journals = load_config_names(resolved_config_path)
    missing_rows = [row for row in directory_rows if row[0] not in known_journals]

    suggestions: list[JournalSuggestion] = []
    for index, (journal_name, publisher) in enumerate(missing_rows, start=1):
        emit_progress(
            progress_callback,
            f"[{index}/{len(missing_rows)}] Searching OpenAlex sources for {journal_name}...",
        )
        candidates = source_searcher(journal_name, publisher)
        suggestions.append(build_suggestion(journal_name, publisher, candidates))

    write_suggestions_sheet(workbook_path, suggestions)
    emit_progress(
        progress_callback,
        f"Wrote {len(suggestions)} suggestion rows to '{SUGGESTIONS_SHEET}'",
    )

    return DiscoverySummary(
        workbook_path=workbook_path,
        config_path=resolved_config_path,
        journals_checked=len(directory_rows),
        missing_journals=len(missing_rows),
        suggestion_rows=len(suggestions),
        output_sheet=SUGGESTIONS_SHEET,
    )
