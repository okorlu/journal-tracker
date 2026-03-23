from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

TRACKER_SETTINGS_SHEET = "Tracker Settings"
TRACK_COLUMN_HEADER = "Track?"
TRACK_TRUE_VALUES = {"1", "true", "yes", "y"}


@dataclass(frozen=True)
class WorkbookSettings:
    workbook_path: Path
    years: int | None = None
    articles_sheet: str | None = None
    directory_sheet: str | None = None
    csv_output_path: Path | None = None
    crossref_mailto: str | None = None
    use_tracked_journals_only: bool = False
    tracked_journals: tuple[str, ...] = ()


def normalize_setting_key(value: str) -> str:
    return " ".join((value or "").strip().lower().split())


def parse_bool(value: Any, default: bool = False) -> bool:
    if value is None:
        return default
    normalized = str(value).strip().lower()
    if not normalized:
        return default
    return normalized in TRACK_TRUE_VALUES


def _resolve_optional_path(base_dir: Path, value: str | None) -> Path | None:
    if not value:
        return None
    path = Path(value).expanduser()
    if not path.is_absolute():
        path = (base_dir / path).resolve()
    return path


def _header_indexes(sheet) -> dict[str, int]:
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    return {
        str(value or "").strip(): index
        for index, value in enumerate(header_row, start=0)
        if str(value or "").strip()
    }


def read_tracked_journals(
    workbook_path: Path,
    directory_sheet_name: str,
) -> tuple[str, ...]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    if directory_sheet_name not in workbook.sheetnames:
        workbook.close()
        return ()

    sheet = workbook[directory_sheet_name]
    header_indexes = _header_indexes(sheet)
    journal_index = header_indexes.get("Journal Name")
    track_index = header_indexes.get(TRACK_COLUMN_HEADER)
    if journal_index is None or track_index is None:
        workbook.close()
        return ()

    tracked: list[str] = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        journal_name = str(row[journal_index] or "").strip()
        if not journal_name:
            continue
        if parse_bool(row[track_index]):
            tracked.append(journal_name)

    workbook.close()
    return tuple(tracked)


def load_workbook_settings(workbook_path: Path) -> WorkbookSettings:
    resolved_path = workbook_path.expanduser().resolve()
    workbook = load_workbook(resolved_path, read_only=True, data_only=True)
    if TRACKER_SETTINGS_SHEET not in workbook.sheetnames:
        workbook.close()
        return WorkbookSettings(workbook_path=resolved_path)

    sheet = workbook[TRACKER_SETTINGS_SHEET]
    values: dict[str, Any] = {}
    for row in sheet.iter_rows(min_row=2, max_col=2, values_only=True):
        key = normalize_setting_key(str(row[0] or ""))
        if key:
            values[key] = row[1]
    workbook.close()

    directory_sheet = str(values.get("journal directory sheet") or "").strip() or None
    tracked_journals = ()
    if parse_bool(values.get("use tracked journals only")):
        tracked_journals = read_tracked_journals(
            resolved_path,
            directory_sheet or "Journal Directory",
        )

    years_value = values.get("years")
    years = int(years_value) if years_value not in (None, "") else None
    csv_output_path = _resolve_optional_path(
        resolved_path.parent,
        str(values.get("csv output") or "").strip() or None,
    )
    crossref_mailto = str(values.get("crossref mailto") or "").strip() or None

    return WorkbookSettings(
        workbook_path=resolved_path,
        years=years,
        articles_sheet=str(values.get("articles sheet") or "").strip() or None,
        directory_sheet=directory_sheet,
        csv_output_path=csv_output_path,
        crossref_mailto=crossref_mailto,
        use_tracked_journals_only=parse_bool(values.get("use tracked journals only")),
        tracked_journals=tracked_journals,
    )
