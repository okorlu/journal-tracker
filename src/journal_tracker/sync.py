from __future__ import annotations

import csv
import json
import re
import shutil
import time
import unicodedata
from copy import copy
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any, Callable, Collection, Iterable
from urllib.error import HTTPError, URLError
from urllib.parse import urlencode, urlparse
from urllib.request import Request, urlopen

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

ARTICLES_SHEET = "Articles"
DIRECTORY_SHEET = "Journal Directory"
META_SHEET = "_openalex_sync_meta"
DEFAULT_PER_PAGE = 200
DEFAULT_YEARS = 3
OPENALEX_API_URL = "https://api.openalex.org/works"
DOI_COLUMN = 7
ARTICLE_URL_COLUMN = 8
CLUSTER_COLUMN = 9
KEY_TOPICS_COLUMN = 10
ADDED_AT_HEADER = "Added At"
DOI_HEADER = "DOI"
ARTICLE_URL_HEADER = "Article URL"
LEGACY_LINK_HEADER = "DOI/Link"
CROSSREF_API_URL = "https://api.crossref.org/works"
ProgressCallback = Callable[[str], None]
DIRECTORY_JOURNAL_HEADER = "Journal Name"
DIRECTORY_PUBLISHER_HEADER = "Publisher"
DIRECTORY_CIRCLE_HEADER = "Circle"
DIRECTORY_CLUSTER_HEADER = "Cluster"
DIRECTORY_QUARTILE_HEADER = "Quartile"
DIRECTORY_WEBSITE_HEADER = "Website"


@dataclass(frozen=True)
class JournalConfig:
    journal_name: str
    source_id: str
    publisher: str | None
    cluster: str | None
    alias: str | None = None


@dataclass(frozen=True)
class JournalDirectoryEntry:
    journal_name: str
    publisher: str
    circle: str
    cluster: str
    quartile: str
    website: str
    source_id: str
    alias: str | None = None


@dataclass(frozen=True)
class JournalSyncResult:
    journal_name: str
    fetched_count: int
    new_count: int
    duplicate_count: int


@dataclass(frozen=True)
class SyncSummary:
    workbook_path: Path
    cutoff_date: date
    journal_results: list[JournalSyncResult]
    total_fetched: int
    total_new_rows: int
    total_duplicates: int
    dry_run: bool
    backup_path: Path | None = None
    workbook_changed: bool = False
    identifier_columns_migrated: bool = False
    added_at_column_added: bool = False
    added_at_backfilled: int = 0


@dataclass(frozen=True)
class WorkbookWriteResult:
    backup_path: Path
    workbook_changed: bool
    identifier_columns_migrated: bool
    added_at_column_added: bool
    added_at_backfilled: int


@dataclass(frozen=True)
class ArticleIdentifiers:
    doi_url: str
    article_url: str


@dataclass(frozen=True)
class CrossrefCandidate:
    doi_url: str
    article_url: str


def default_config_path() -> Path:
    return Path(__file__).resolve().parents[2] / "config" / "openalex_sources.json"


def load_config(config_path: Path) -> dict[str, JournalConfig]:
    items = json.loads(config_path.read_text(encoding="utf-8"))
    config: dict[str, JournalConfig] = {}
    for item in items:
        config[item["journal_name"]] = JournalConfig(
            journal_name=item["journal_name"],
            source_id=item["source_id"],
            publisher=item.get("publisher"),
            cluster=item.get("cluster"),
            alias=item.get("alias"),
        )
    return config


def emit_progress(progress_callback: ProgressCallback | None, message: str) -> None:
    if progress_callback:
        progress_callback(message)


def json_get(url: str, user_agent: str) -> dict[str, Any]:
    request = Request(
        url,
        headers={
            "User-Agent": user_agent,
            "Accept": "application/json",
        },
    )
    with urlopen(request, timeout=60) as response:
        return json.load(response)


def openalex_get(url: str) -> dict[str, Any]:
    return json_get(url, "journal-tracker-openalex-sync/1.0")


def crossref_get(url: str, mailto: str | None = None) -> dict[str, Any]:
    user_agent = "journal-tracker-crossref-enrichment/1.0"
    if mailto:
        user_agent = f"{user_agent} (mailto:{mailto})"
    return json_get(url, user_agent)


def rolling_cutoff(today: date, years: int) -> date:
    try:
        return today.replace(year=today.year - years)
    except ValueError:
        return today.replace(month=2, day=28, year=today.year - years)


def normalize_text(value: str) -> str:
    collapsed = " ".join((value or "").strip().split())
    ascii_value = unicodedata.normalize("NFKD", collapsed)
    ascii_value = "".join(ch for ch in ascii_value if not unicodedata.combining(ch))
    ascii_value = re.sub(r"[^0-9a-zA-Z]+", " ", ascii_value)
    return " ".join(ascii_value.lower().split())


def normalize_doi(value: str | None) -> str | None:
    if not value:
        return None
    match = re.search(r"10\.\S+", value, re.IGNORECASE)
    if not match:
        return None
    doi = match.group(0).strip().rstrip(" .;,)")
    return doi.lower()


def read_directory_sheet(
    workbook_path: Path,
    config: dict[str, JournalConfig],
    directory_sheet_name: str = DIRECTORY_SHEET,
    selected_journals: Collection[str] | None = None,
) -> list[JournalDirectoryEntry]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    if directory_sheet_name not in workbook.sheetnames:
        raise ValueError(f"Workbook is missing the '{directory_sheet_name}' sheet.")

    sheet = workbook[directory_sheet_name]
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    header_indexes = {
        str(value or "").strip(): index
        for index, value in enumerate(header_row)
        if str(value or "").strip()
    }
    journal_index = header_indexes.get(DIRECTORY_JOURNAL_HEADER)
    if journal_index is None:
        workbook.close()
        raise ValueError(
            f"'{directory_sheet_name}' must include a '{DIRECTORY_JOURNAL_HEADER}' column."
        )

    selected_lookup = None if selected_journals is None else set(selected_journals)
    seen_selected: set[str] = set()
    entries: list[JournalDirectoryEntry] = []

    def row_value(row_values: tuple[Any, ...], header_name: str) -> Any:
        index = header_indexes.get(header_name)
        if index is None or index >= len(row_values):
            return None
        return row_values[index]

    for row in sheet.iter_rows(min_row=2, values_only=True):
        journal_name = str(row[journal_index] or "").strip()
        if not journal_name:
            continue
        if selected_lookup is not None and journal_name not in selected_lookup:
            continue
        if journal_name not in config:
            workbook.close()
            raise ValueError(
                f"Journal '{journal_name}' is present in the workbook but missing from config."
            )
        source = config[journal_name]
        seen_selected.add(journal_name)
        entries.append(
            JournalDirectoryEntry(
                journal_name=journal_name,
                publisher=row_value(row, DIRECTORY_PUBLISHER_HEADER) or source.publisher or "",
                circle=row_value(row, DIRECTORY_CIRCLE_HEADER) or "",
                cluster=row_value(row, DIRECTORY_CLUSTER_HEADER) or source.cluster or "",
                quartile=row_value(row, DIRECTORY_QUARTILE_HEADER) or "",
                website=row_value(row, DIRECTORY_WEBSITE_HEADER) or "",
                source_id=source.source_id,
                alias=source.alias,
            )
        )
    workbook.close()
    if selected_lookup is not None:
        missing = [name for name in selected_journals or () if name not in seen_selected]
        if missing:
            quoted = ", ".join(f"'{name}'" for name in missing)
            raise ValueError(
                f"The following journals were requested but not found in "
                f"'{directory_sheet_name}': {quoted}"
            )
    return entries


def fetch_works(
    source_id: str,
    cutoff_date: date,
    api_key: str,
    progress_callback: ProgressCallback | None = None,
    progress_label: str | None = None,
) -> list[dict[str, Any]]:
    results: list[dict[str, Any]] = []
    cursor = "*"
    page_count = 0
    while True:
        page_count += 1
        params = {
            "filter": (
                "primary_location.source.id:"
                f"{source_id},from_publication_date:{cutoff_date.isoformat()}"
            ),
            "per-page": str(DEFAULT_PER_PAGE),
            "cursor": cursor,
            "api_key": api_key,
        }
        url = f"{OPENALEX_API_URL}?{urlencode(params)}"
        try:
            payload = openalex_get(url)
        except HTTPError as exc:
            body = exc.read().decode("utf-8", errors="replace")
            raise RuntimeError(
                f"OpenAlex request failed for {source_id}: HTTP {exc.code} {body}"
            ) from exc
        except URLError as exc:
            raise RuntimeError(f"OpenAlex request failed for {source_id}: {exc}") from exc

        results.extend(payload.get("results", []))
        cursor = payload.get("meta", {}).get("next_cursor")
        if cursor:
            label = progress_label or source_id
            emit_progress(
                progress_callback,
                f"  {label}: page {page_count} fetched, {len(results)} records so far",
            )
        if not cursor:
            break
        time.sleep(0.1)
    return results


def format_authors(authorships: Iterable[dict[str, Any]]) -> str:
    names: list[str] = []
    for authorship in authorships:
        author = authorship.get("author") or {}
        name = (author.get("display_name") or authorship.get("raw_author_name") or "").strip()
        if name and name not in names:
            names.append(name)
    return "; ".join(names)


def format_volume_issue(biblio: dict[str, Any]) -> str:
    volume = biblio.get("volume")
    issue = biblio.get("issue")
    parts: list[str] = []
    if volume:
        parts.append(f"Vol. {volume}")
    if issue:
        parts.append(f"No. {issue}")
    return ", ".join(parts)


def format_pages(biblio: dict[str, Any]) -> str:
    first_page = biblio.get("first_page")
    last_page = biblio.get("last_page")
    if first_page and last_page:
        return f"{first_page}-{last_page}"
    return first_page or last_page or ""


def format_topics(work: dict[str, Any]) -> str:
    items: list[str] = []
    source_items = (work.get("keywords") or []) or (work.get("topics") or [])
    for item in source_items:
        label = (item.get("display_name") or "").strip()
        if label and label not in items:
            items.append(label)
        if len(items) == 5:
            break
    return ", ".join(items)


def doi_url(value: str | None) -> str:
    normalized = normalize_doi(value)
    if not normalized:
        return ""
    return f"https://doi.org/{normalized}"


def is_probably_article_url(value: str | None) -> bool:
    if not value:
        return False

    parsed = urlparse(value)
    if parsed.scheme not in {"http", "https"} or not parsed.netloc:
        return False

    path = parsed.path.lower().rstrip("/")
    host = parsed.netloc.lower()

    blocked_fragments = (
        "/toc/",
        "/current",
        "/latest-issue",
        "/issues",
        "/journal/",
    )
    if any(fragment in path for fragment in blocked_fragments):
        return False

    if "tandfonline.com" in host and "/doi/" not in path:
        return False
    if "cambridge.org" in host and "/article/" not in path:
        return False
    if "researchgate.net" in host and "/journal/" in path:
        return False
    if path in {"", "/"}:
        return False

    return True


def first_author_last_name(authorships: Iterable[dict[str, Any]]) -> str:
    for authorship in authorships:
        author = authorship.get("author") or {}
        name = (author.get("display_name") or authorship.get("raw_author_name") or "").strip()
        if name:
            return normalize_text(name.split()[-1])
    return ""


def crossref_year(item: dict[str, Any]) -> int | None:
    for key in ("published-print", "published-online", "issued"):
        date_parts = (item.get(key) or {}).get("date-parts") or []
        if date_parts and date_parts[0]:
            return int(date_parts[0][0])
    return None


def crossref_first_author_last_name(item: dict[str, Any]) -> str:
    authors = item.get("author") or []
    if not authors:
        return ""
    author = authors[0] or {}
    family = (author.get("family") or "").strip()
    if family:
        return normalize_text(family)
    given = (author.get("given") or "").strip()
    if given:
        return normalize_text(given.split()[-1])
    return ""


def lookup_crossref_candidate(
    work: dict[str, Any],
    directory_entry: JournalDirectoryEntry,
    mailto: str | None = None,
    getter: Callable[[str, str | None], dict[str, Any]] = crossref_get,
) -> CrossrefCandidate | None:
    title = (work.get("display_name") or "").strip()
    if not title:
        return None

    query_parts = [title]
    author_last_name = first_author_last_name(work.get("authorships") or [])
    if author_last_name:
        query_parts.append(author_last_name)
    publication_year = work.get("publication_year")
    if publication_year:
        query_parts.append(str(publication_year))

    params = {
        "rows": "5",
        "query.bibliographic": " ".join(query_parts),
        "select": (
            "DOI,URL,title,container-title,author,published-print,published-online,issued,score"
        ),
    }
    url = f"{CROSSREF_API_URL}?{urlencode(params)}"

    try:
        payload = getter(url, mailto)
    except Exception:
        return None

    target_title = normalize_text(title)
    expected_journals = {
        normalize_text(directory_entry.journal_name),
        normalize_text(directory_entry.alias or ""),
    }
    expected_journals.discard("")
    target_year = int(publication_year) if publication_year else None
    target_author = author_last_name

    for item in (payload.get("message") or {}).get("items", []):
        candidate_titles = item.get("title") or []
        if not candidate_titles:
            continue
        if normalize_text(candidate_titles[0]) != target_title:
            continue

        container_titles = {
            normalize_text(value) for value in (item.get("container-title") or []) if value
        }
        if (
            expected_journals
            and container_titles
            and not container_titles.intersection(expected_journals)
        ):
            continue

        candidate_year = crossref_year(item)
        if target_year and candidate_year and candidate_year != target_year:
            continue

        candidate_author = crossref_first_author_last_name(item)
        if target_author and candidate_author and target_author != candidate_author:
            continue

        return CrossrefCandidate(
            doi_url=doi_url(item.get("DOI")),
            article_url=item.get("URL") or "",
        )

    return None


def resolve_article_identifiers(
    work: dict[str, Any],
    directory_entry: JournalDirectoryEntry,
    crossref_lookup: Callable[[dict[str, Any], JournalDirectoryEntry], CrossrefCandidate | None]
    | None = None,
) -> ArticleIdentifiers:
    resolved_doi_url = doi_url(work.get("doi"))
    primary_location = work.get("primary_location") or {}
    article_url = ""
    landing_page_url = primary_location.get("landing_page_url")
    if is_probably_article_url(landing_page_url):
        article_url = landing_page_url

    if crossref_lookup and (not resolved_doi_url or not article_url):
        candidate = crossref_lookup(work, directory_entry)
        if candidate:
            if not resolved_doi_url and candidate.doi_url:
                resolved_doi_url = candidate.doi_url
            if not article_url and is_probably_article_url(candidate.article_url):
                article_url = candidate.article_url

    if not article_url:
        article_url = work.get("id") or ""

    return ArticleIdentifiers(
        doi_url=resolved_doi_url,
        article_url=article_url,
    )


def normalized_title_key(title: str, journal: str, year: Any) -> str:
    return "|".join(
        [
            normalize_text(title),
            normalize_text(journal),
            normalize_text(str(year or "")),
        ]
    )


def ensure_meta_sheet(workbook) -> Any:
    if META_SHEET in workbook.sheetnames:
        return workbook[META_SHEET]

    meta_sheet = workbook.create_sheet(META_SHEET)
    meta_sheet.sheet_state = "hidden"
    meta_sheet.append(["work_id", "doi", "normalized_key", "synced_at"])
    return meta_sheet


def clone_cell_style(source, target) -> None:
    if source.has_style:
        target._style = copy(source._style)
    if source.number_format:
        target.number_format = source.number_format
    if source.font:
        target.font = copy(source.font)
    if source.fill:
        target.fill = copy(source.fill)
    if source.border:
        target.border = copy(source.border)
    if source.alignment:
        target.alignment = copy(source.alignment)
    if source.protection:
        target.protection = copy(source.protection)


def set_cell_link(cell, value: str) -> None:
    cell.value = value
    cell.hyperlink = value if value else None


def ensure_identifier_columns(articles_sheet) -> tuple[int, int, bool]:
    headers = [
        (articles_sheet.cell(row=1, column=column_index).value or "").strip()
        for column_index in range(1, articles_sheet.max_column + 1)
    ]
    if DOI_HEADER in headers and ARTICLE_URL_HEADER in headers:
        return headers.index(DOI_HEADER) + 1, headers.index(ARTICLE_URL_HEADER) + 1, False

    if LEGACY_LINK_HEADER not in headers:
        raise ValueError(
            f"Articles sheet must include '{DOI_HEADER}' and '{ARTICLE_URL_HEADER}' "
            f"or legacy '{LEGACY_LINK_HEADER}'."
        )

    legacy_column = headers.index(LEGACY_LINK_HEADER) + 1
    articles_sheet.insert_cols(legacy_column + 1)
    articles_sheet.cell(row=1, column=legacy_column).value = DOI_HEADER
    article_url_header_cell = articles_sheet.cell(row=1, column=legacy_column + 1)
    clone_cell_style(articles_sheet.cell(row=1, column=legacy_column), article_url_header_cell)
    article_url_header_cell.value = ARTICLE_URL_HEADER

    source_letter = get_column_letter(legacy_column)
    target_letter = get_column_letter(legacy_column + 1)
    source_width = articles_sheet.column_dimensions[source_letter].width
    if source_width is not None:
        articles_sheet.column_dimensions[target_letter].width = source_width

    for row_index in range(2, articles_sheet.max_row + 1):
        doi_cell = articles_sheet.cell(row=row_index, column=legacy_column)
        article_url_cell = articles_sheet.cell(row=row_index, column=legacy_column + 1)
        clone_cell_style(doi_cell, article_url_cell)

        legacy_value = str(doi_cell.value or "").strip()
        legacy_link = doi_cell.hyperlink.target if doi_cell.hyperlink else legacy_value
        normalized = normalize_doi(legacy_value or legacy_link)
        if normalized:
            set_cell_link(doi_cell, doi_url(normalized))
            set_cell_link(article_url_cell, "")
        else:
            set_cell_link(doi_cell, "")
            set_cell_link(article_url_cell, legacy_link if legacy_link else legacy_value)

    return legacy_column, legacy_column + 1, True


def ensure_added_at_column(articles_sheet) -> tuple[int, bool]:
    headers = [
        (articles_sheet.cell(row=1, column=column_index).value or "").strip()
        for column_index in range(1, articles_sheet.max_column + 1)
    ]
    for column_index, header in enumerate(headers, start=1):
        if header == ADDED_AT_HEADER:
            return column_index, False

    added_at_column = articles_sheet.max_column + 1
    source_column = max(1, added_at_column - 1)
    for row_index in range(1, articles_sheet.max_row + 1):
        source = articles_sheet.cell(row=row_index, column=source_column)
        target = articles_sheet.cell(row=row_index, column=added_at_column)
        clone_cell_style(source, target)
    articles_sheet.cell(row=1, column=added_at_column).value = ADDED_AT_HEADER

    source_letter = get_column_letter(source_column)
    target_letter = get_column_letter(added_at_column)
    source_width = articles_sheet.column_dimensions[source_letter].width
    if source_width is not None:
        articles_sheet.column_dimensions[target_letter].width = source_width

    return added_at_column, True


def backfill_added_at_from_meta(articles_sheet, meta_sheet, added_at_column: int) -> int:
    synced_at_by_key: dict[str, str] = {}
    for work_id, doi, normalized_key, synced_at in meta_sheet.iter_rows(
        min_row=2,
        max_col=4,
        values_only=True,
    ):
        del work_id, doi
        if normalized_key and synced_at and normalized_key not in synced_at_by_key:
            synced_at_by_key[normalized_key] = synced_at

    updated_count = 0
    for row_index in range(2, articles_sheet.max_row + 1):
        added_at_cell = articles_sheet.cell(row=row_index, column=added_at_column)
        if added_at_cell.value:
            continue
        normalized_key = normalized_title_key(
            str(articles_sheet.cell(row=row_index, column=1).value or ""),
            str(articles_sheet.cell(row=row_index, column=3).value or ""),
            articles_sheet.cell(row=row_index, column=5).value or "",
        )
        synced_at = synced_at_by_key.get(normalized_key)
        if synced_at:
            added_at_cell.value = synced_at
            updated_count += 1

    return updated_count


def prepare_articles_sheet(workbook, articles_sheet_name: str):
    if articles_sheet_name not in workbook.sheetnames:
        raise ValueError(f"Workbook is missing the '{articles_sheet_name}' sheet.")

    articles_sheet = workbook[articles_sheet_name]
    meta_sheet = ensure_meta_sheet(workbook)
    doi_column, article_url_column, identifier_columns_migrated = ensure_identifier_columns(
        articles_sheet
    )
    added_at_column, added_at_column_added = ensure_added_at_column(articles_sheet)
    added_at_backfilled = backfill_added_at_from_meta(
        articles_sheet,
        meta_sheet,
        added_at_column,
    )
    return (
        articles_sheet,
        meta_sheet,
        doi_column,
        article_url_column,
        added_at_column,
        identifier_columns_migrated,
        added_at_column_added,
        added_at_backfilled,
    )


def read_existing_indexes(articles_sheet, meta_sheet) -> tuple[set[str], set[str], set[str]]:
    doi_index: set[str] = set()
    work_id_index: set[str] = set()
    normalized_index: set[str] = set()

    for row in articles_sheet.iter_rows(min_row=2, max_col=7):
        title = row[0].value or ""
        journal = row[2].value or ""
        year = row[4].value or ""
        doi_value = row[DOI_COLUMN - 1].value or ""
        doi = normalize_doi(doi_value)
        if doi:
            doi_index.add(doi)
        normalized_index.add(normalized_title_key(str(title), str(journal), year))

    for row in meta_sheet.iter_rows(min_row=2, max_col=3, values_only=True):
        work_id = row[0] or ""
        doi = normalize_doi(row[1] or "")
        normalized_key = row[2] or ""
        if work_id:
            work_id_index.add(work_id)
        if doi:
            doi_index.add(doi)
        if normalized_key:
            normalized_index.add(normalized_key)

    return doi_index, work_id_index, normalized_index


def build_rows(
    directory_entry: JournalDirectoryEntry,
    works: list[dict[str, Any]],
    doi_index: set[str],
    work_id_index: set[str],
    normalized_index: set[str],
    crossref_lookup: Callable[[dict[str, Any], JournalDirectoryEntry], CrossrefCandidate | None]
    | None = None,
) -> tuple[list[list[Any]], list[tuple[str, str | None, str]], int]:
    rows: list[list[Any]] = []
    meta_rows: list[tuple[str, str | None, str]] = []
    duplicate_count = 0

    sorted_works = sorted(
        works,
        key=lambda work: (
            work.get("publication_date") or "",
            work.get("publication_year") or 0,
            (work.get("display_name") or "").lower(),
        ),
    )

    for work in sorted_works:
        work_id = work.get("id") or ""
        doi = normalize_doi(work.get("doi"))
        normalized_key = normalized_title_key(
            work.get("display_name") or "",
            directory_entry.journal_name,
            work.get("publication_year") or "",
        )

        if doi and doi in doi_index:
            duplicate_count += 1
            continue
        if work_id and work_id in work_id_index:
            duplicate_count += 1
            continue
        if normalized_key in normalized_index:
            duplicate_count += 1
            continue

        biblio = work.get("biblio") or {}
        identifiers = resolve_article_identifiers(
            work,
            directory_entry,
            crossref_lookup=crossref_lookup,
        )
        rows.append(
            [
                work.get("display_name") or "",
                format_authors(work.get("authorships") or []),
                directory_entry.journal_name,
                format_volume_issue(biblio),
                work.get("publication_year") or "",
                format_pages(biblio),
                identifiers.doi_url,
                identifiers.article_url,
                directory_entry.cluster,
                format_topics(work),
            ]
        )
        meta_rows.append((work_id, doi, normalized_key))
        if doi:
            doi_index.add(doi)
        if work_id:
            work_id_index.add(work_id)
        normalized_index.add(normalized_key)

    return rows, meta_rows, duplicate_count


def clone_row_style(sheet, template_row: int, target_row: int, total_columns: int) -> None:
    for column_index in range(1, total_columns + 1):
        source = sheet.cell(row=template_row, column=column_index)
        target = sheet.cell(row=target_row, column=column_index)
        clone_cell_style(source, target)
    if sheet.row_dimensions[template_row].height is not None:
        sheet.row_dimensions[target_row].height = sheet.row_dimensions[template_row].height


def row_has_data(sheet, row_index: int, total_columns: int) -> bool:
    for column_index in range(1, total_columns + 1):
        if sheet.cell(row=row_index, column=column_index).value not in (None, ""):
            return True
    return False


def append_rows(
    workbook_path: Path,
    rows: list[list[Any]],
    meta_rows: list[tuple[str, str | None, str]],
    articles_sheet_name: str = ARTICLES_SHEET,
) -> WorkbookWriteResult:
    workbook = load_workbook(workbook_path)
    (
        articles_sheet,
        meta_sheet,
        doi_column,
        article_url_column,
        added_at_column,
        identifier_columns_migrated,
        added_at_column_added,
        added_at_backfilled,
    ) = prepare_articles_sheet(workbook, articles_sheet_name)
    template_row = 2 if articles_sheet.max_row >= 2 else 1
    synced_at = datetime.now().isoformat(timespec="seconds")
    workbook_changed = (
        identifier_columns_migrated
        or added_at_column_added
        or added_at_backfilled > 0
        or bool(rows)
    )

    for values, meta in zip(rows, meta_rows):
        if articles_sheet.max_row >= 2 and not row_has_data(
            articles_sheet, 2, articles_sheet.max_column
        ):
            next_row = 2
        else:
            next_row = articles_sheet.max_row + 1
        clone_row_style(articles_sheet, template_row, next_row, articles_sheet.max_column)
        for column_index, value in enumerate(values, start=1):
            cell = articles_sheet.cell(row=next_row, column=column_index)
            cell.value = value
            if column_index in {doi_column, article_url_column} and value:
                cell.hyperlink = value
        articles_sheet.cell(row=next_row, column=added_at_column).value = synced_at
        meta_sheet.append([meta[0], meta[1], meta[2], synced_at])

    backup_path = workbook_path.with_name(
        f"{workbook_path.stem}.{datetime.now().strftime('%Y%m%d-%H%M%S')}.bak{workbook_path.suffix}"
    )
    shutil.copy2(workbook_path, backup_path)
    workbook.save(workbook_path)
    workbook.close()
    return WorkbookWriteResult(
        backup_path=backup_path,
        workbook_changed=workbook_changed,
        identifier_columns_migrated=identifier_columns_migrated,
        added_at_column_added=added_at_column_added,
        added_at_backfilled=added_at_backfilled,
    )


def export_articles_to_csv(
    workbook_path: Path,
    csv_path: Path,
    articles_sheet_name: str = ARTICLES_SHEET,
) -> Path:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    if articles_sheet_name not in workbook.sheetnames:
        raise ValueError(f"Workbook is missing the '{articles_sheet_name}' sheet.")

    csv_path = csv_path.expanduser().resolve()
    csv_path.parent.mkdir(parents=True, exist_ok=True)

    articles_sheet = workbook[articles_sheet_name]
    with csv_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle)
        rows = articles_sheet.iter_rows(values_only=True)
        header = next(rows, None)
        if header is None:
            workbook.close()
            return csv_path

        header_values = ["" if value is None else value for value in header]
        if LEGACY_LINK_HEADER in header_values:
            legacy_index = header_values.index(LEGACY_LINK_HEADER)
            header_values = (
                header_values[:legacy_index]
                + [DOI_HEADER, ARTICLE_URL_HEADER]
                + header_values[legacy_index + 1 :]
            )
        has_added_at = ADDED_AT_HEADER in header_values
        if not has_added_at:
            header_values.append(ADDED_AT_HEADER)
        writer.writerow(header_values)

        for row in rows:
            values = ["" if value is None else value for value in row]
            if LEGACY_LINK_HEADER in ["" if value is None else value for value in header]:
                legacy_index = ["" if value is None else value for value in header].index(
                    LEGACY_LINK_HEADER
                )
                legacy_value = values[legacy_index]
                normalized = normalize_doi(legacy_value)
                if normalized:
                    values = (
                        values[:legacy_index]
                        + [doi_url(normalized), ""]
                        + values[legacy_index + 1 :]
                    )
                else:
                    values = values[:legacy_index] + ["", legacy_value] + values[legacy_index + 1 :]
            if not has_added_at:
                values.append("")
            writer.writerow(values)

    workbook.close()
    return csv_path


def sync_workbook(
    workbook_path: Path,
    config_path: Path,
    api_key: str,
    years: int = DEFAULT_YEARS,
    dry_run: bool = False,
    articles_sheet: str = ARTICLES_SHEET,
    directory_sheet: str = DIRECTORY_SHEET,
    journal_names: Collection[str] | None = None,
    today: date | None = None,
    fetcher: Callable[[str, date, str], list[dict[str, Any]]] = fetch_works,
    progress_callback: ProgressCallback | None = None,
    crossref_mailto: str | None = None,
    crossref_lookup: Callable[[dict[str, Any], JournalDirectoryEntry], CrossrefCandidate | None]
    | None = None,
) -> SyncSummary:
    workbook_path = workbook_path.expanduser().resolve()
    config_path = config_path.expanduser().resolve()
    if not workbook_path.exists():
        raise FileNotFoundError(f"Workbook not found: {workbook_path}")
    if not config_path.exists():
        raise FileNotFoundError(f"Config not found: {config_path}")

    emit_progress(progress_callback, f"Loading journal config from {config_path}")
    config = load_config(config_path)
    emit_progress(progress_callback, f"Reading journal directory from {workbook_path}")
    directory_entries = read_directory_sheet(
        workbook_path,
        config,
        directory_sheet_name=directory_sheet,
        selected_journals=journal_names,
    )
    cutoff_date = rolling_cutoff(today or date.today(), years)
    emit_progress(
        progress_callback,
        f"Loaded {len(directory_entries)} journals. Cutoff date: {cutoff_date.isoformat()}",
    )

    workbook = load_workbook(workbook_path)
    (
        articles_sheet_ref,
        meta_sheet,
        doi_column,
        article_url_column,
        added_at_column,
        identifier_columns_migrated,
        added_at_column_added,
        added_at_backfilled,
    ) = prepare_articles_sheet(workbook, articles_sheet)
    doi_index, work_id_index, normalized_index = read_existing_indexes(
        articles_sheet_ref, meta_sheet
    )
    workbook.close()
    del doi_column, article_url_column, added_at_column

    journal_results: list[JournalSyncResult] = []
    all_new_rows: list[list[Any]] = []
    all_meta_rows: list[tuple[str, str | None, str]] = []
    total_fetched = 0
    total_duplicates = 0
    if crossref_lookup is None:
        crossref_cache: dict[str, CrossrefCandidate | None] = {}

        def crossref_lookup(
            work: dict[str, Any],
            directory_entry: JournalDirectoryEntry,
        ) -> CrossrefCandidate | None:
            cache_key = normalized_title_key(
                work.get("display_name") or "",
                directory_entry.journal_name,
                work.get("publication_year") or "",
            )
            if cache_key not in crossref_cache:
                crossref_cache[cache_key] = lookup_crossref_candidate(
                    work,
                    directory_entry,
                    mailto=crossref_mailto,
                )
            return crossref_cache[cache_key]

    for index, entry in enumerate(directory_entries, start=1):
        emit_progress(
            progress_callback,
            f"[{index}/{len(directory_entries)}] Fetching {entry.journal_name}...",
        )
        if fetcher is fetch_works:
            works = fetcher(
                entry.source_id,
                cutoff_date,
                api_key,
                progress_callback=progress_callback,
                progress_label=entry.journal_name,
            )
        else:
            works = fetcher(entry.source_id, cutoff_date, api_key)
        rows, meta_rows, duplicate_count = build_rows(
            entry,
            works,
            doi_index,
            work_id_index,
            normalized_index,
            crossref_lookup=crossref_lookup,
        )
        emit_progress(
            progress_callback,
            f"[{index}/{len(directory_entries)}] {entry.journal_name}: "
            f"fetched={len(works)} new={len(rows)} duplicates={duplicate_count}",
        )
        journal_results.append(
            JournalSyncResult(
                journal_name=entry.journal_name,
                fetched_count=len(works),
                new_count=len(rows),
                duplicate_count=duplicate_count,
            )
        )
        total_fetched += len(works)
        total_duplicates += duplicate_count
        all_new_rows.extend(rows)
        all_meta_rows.extend(meta_rows)

    backup_path = None
    workbook_changed = (
        identifier_columns_migrated or added_at_column_added or added_at_backfilled > 0
    )
    if not dry_run and (all_new_rows or workbook_changed):
        emit_progress(
            progress_callback,
            f"Writing workbook updates ({len(all_new_rows)} new rows)...",
        )
        write_result = append_rows(
            workbook_path,
            all_new_rows,
            all_meta_rows,
            articles_sheet_name=articles_sheet,
        )
        backup_path = write_result.backup_path
        workbook_changed = write_result.workbook_changed
        identifier_columns_migrated = write_result.identifier_columns_migrated
        added_at_column_added = write_result.added_at_column_added
        added_at_backfilled = write_result.added_at_backfilled

    return SyncSummary(
        workbook_path=workbook_path,
        cutoff_date=cutoff_date,
        journal_results=journal_results,
        total_fetched=total_fetched,
        total_new_rows=len(all_new_rows),
        total_duplicates=total_duplicates,
        dry_run=dry_run,
        backup_path=backup_path,
        workbook_changed=workbook_changed,
        identifier_columns_migrated=identifier_columns_migrated,
        added_at_column_added=added_at_column_added,
        added_at_backfilled=added_at_backfilled,
    )
